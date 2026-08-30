"""
SMART-SUP — Import du Catalogue de supervision dans le référentiel
===================================================================

Alimente les tables `domaines` et `services` à partir du classeur
« Catalogue_Services_SupervisionV4.xlsx » (onglet « Catalogue »).

Pourquoi c'est important : dans le système actuel, le service impacté est
saisi en texte libre dans chaque mail. Résultat, il est impossible d'agréger
de façon fiable un reporting par service. En rattachant l'incident à une
entrée du catalogue, on obtient trois gains d'un coup :

  1. la priorité de l'incident se pré-remplit depuis le service (le catalogue
     porte déjà `priorite_defaut`) — moins de saisie, moins d'incohérences ;
  2. le libellé du service devient stable d'un mail à l'autre ;
  3. les rapports hebdomadaire et mensuel peuvent enfin compter par service
     et par domaine sans dédoublonnage manuel.

Import idempotent : relancer le script met à jour les services existants
sans créer de doublons.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from db.schema_v5 import get_conn, init_db  # noqa: E402

# Colonnes de l'onglet « Catalogue » (les données commencent ligne 5)
COL_DOMAINE, COL_SERVICE, COL_PRIORITE, COL_SUPERVISE = 0, 1, 2, 3
COL_OUTIL, COL_SIM, COL_SONDE = 4, 5, 6
COL_OBSERVATIONS, COL_ORIGINE = 8, 9

PREMIERE_LIGNE = 5


def _txt(valeur) -> str | None:
    """Normalise une cellule en texte propre, ou None si vide."""
    if valeur is None:
        return None
    texte = str(valeur).strip()
    return texte or None


def importer_catalogue(chemin_xlsx: Path | str, db_path: Path | str | None = None) -> dict:
    """Importe le catalogue et renvoie un bilan chiffré."""
    init_db(db_path) if db_path else init_db()

    wb = load_workbook(str(chemin_xlsx), read_only=True, data_only=True)
    ws = wb["Catalogue"]

    lignes = [
        r for r in ws.iter_rows(min_row=PREMIERE_LIGNE, values_only=True)
        if r and _txt(r[COL_SERVICE])
    ]

    bilan = {
        "lignes_lues": len(lignes),
        "domaines": 0,
        "services_crees": 0,
        "services_maj": 0,
        "sans_priorite": 0,
        "non_supervises": 0,
        "p1_non_supervises": [],
    }

    conn = get_conn(db_path) if db_path else get_conn()
    try:
        with conn:
            for ligne in lignes:
                domaine = _txt(ligne[COL_DOMAINE]) or "Non classé"
                service = _txt(ligne[COL_SERVICE])
                priorite = _txt(ligne[COL_PRIORITE])
                supervise_txt = (_txt(ligne[COL_SUPERVISE]) or "").lower()
                supervise = 1 if supervise_txt.startswith("oui") else 0

                # --- Domaine (créé à la volée) ---
                conn.execute(
                    "INSERT OR IGNORE INTO domaines (nom) VALUES (?)", (domaine,)
                )
                domaine_id = conn.execute(
                    "SELECT id FROM domaines WHERE nom = ?", (domaine,)
                ).fetchone()["id"]

                # --- Service (upsert) ---
                existant = conn.execute(
                    "SELECT id FROM services WHERE domaine_id = ? AND nom = ?",
                    (domaine_id, service),
                ).fetchone()

                valeurs = (
                    priorite,
                    supervise,
                    _txt(ligne[COL_OUTIL]),
                    _txt(ligne[COL_SIM]),
                    _txt(ligne[COL_SONDE]),
                    _txt(ligne[COL_OBSERVATIONS]),
                    _txt(ligne[COL_ORIGINE]),
                )

                if existant:
                    conn.execute(
                        """UPDATE services SET priorite_defaut = ?, supervise = ?,
                           outil_supervision = ?, sim_msisdn_test = ?,
                           sonde_terminal = ?, observations = ?, origine = ?
                           WHERE id = ?""",
                        (*valeurs, existant["id"]),
                    )
                    bilan["services_maj"] += 1
                else:
                    conn.execute(
                        """INSERT INTO services
                           (domaine_id, nom, priorite_defaut, supervise,
                            outil_supervision, sim_msisdn_test, sonde_terminal,
                            observations, origine)
                           VALUES (?,?,?,?,?,?,?,?,?)""",
                        (domaine_id, service, *valeurs),
                    )
                    bilan["services_crees"] += 1

                # --- Signaux qualité ---
                if not priorite:
                    bilan["sans_priorite"] += 1
                if not supervise:
                    bilan["non_supervises"] += 1
                    if priorite == "P1":
                        bilan["p1_non_supervises"].append(f"{domaine} / {service}")

        bilan["domaines"] = conn.execute(
            "SELECT COUNT(*) AS n FROM domaines"
        ).fetchone()["n"]
    finally:
        conn.close()

    wb.close()
    return bilan


def creer_equipes_depuis_domaines(db_path: Path | str | None = None) -> int:
    """
    Crée une équipe de notification par domaine métier.

    Point de départ raisonnable, à ajuster ensuite dans l'administration :
    les vraies équipes ne recoupent pas forcément les domaines un pour un.
    """
    conn = get_conn(db_path) if db_path else get_conn()
    crees = 0
    try:
        with conn:
            for row in conn.execute("SELECT id, nom FROM domaines ORDER BY nom"):
                code = (
                    row["nom"].upper()
                    .replace(" ", "_").replace("(", "").replace(")", "")
                    .replace("/", "_").replace("é", "E").replace("è", "E")
                )[:40]
                cur = conn.execute(
                    """INSERT OR IGNORE INTO equipes (code, nom, domaine_id)
                       VALUES (?,?,?)""",
                    (code, row["nom"], row["id"]),
                )
                crees += cur.rowcount
    finally:
        conn.close()
    return crees


if __name__ == "__main__":
    chemin = sys.argv[1] if len(sys.argv) > 1 else (
        "/mnt/user-data/uploads/Catalogue_Services_SupervisionV4.xlsx"
    )
    resultat = importer_catalogue(chemin)
    nb_equipes = creer_equipes_depuis_domaines()

    print("=" * 66)
    print("IMPORT DU CATALOGUE DE SUPERVISION")
    print("=" * 66)
    print(f"  Lignes lues            : {resultat['lignes_lues']}")
    print(f"  Domaines               : {resultat['domaines']}")
    print(f"  Services créés         : {resultat['services_crees']}")
    print(f"  Services mis à jour    : {resultat['services_maj']}")
    print(f"  Équipes créées         : {nb_equipes}")
    print()
    print("  SIGNAUX QUALITÉ")
    print(f"  Services sans priorité : {resultat['sans_priorite']}")
    print(f"  Services non supervisés: {resultat['non_supervises']}")
    if resultat["p1_non_supervises"]:
        print(f"  ALERTE — P1 non supervisés ({len(resultat['p1_non_supervises'])}) :")
        for item in resultat["p1_non_supervises"]:
            print(f"      - {item}")
