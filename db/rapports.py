"""
SMART-SUP — Rapports
=====================

Trois rapports demandés au cahier des charges (§7), tous construits depuis les
mêmes données : `incidents` + `incident_communications`. Aucune ressaisie.

  · fin de shift   — lecture rapide par l'équipe suivante
  · hebdomadaire   — reprend la structure du classeur Excel existant
  · mensuel        — vision globale, tendances, qualité de service

Le découpage des shifts et le premier jour de semaine sont paramétrables
(Administration → Rapports), donc aucune de ces valeurs n'est écrite ici.

Note sur les durées : la durée mesurée est celle **entre la déclaration et la
clôture documentées chez nous**. Ce n'est pas un indicateur de SLA — la mesure
officielle reste celle de l'outil de ticketing du groupe.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta

from db.parametres import tous_parametres
from db.schema_v5 import get_conn, lire_attributs

# Colonnes du classeur de suivi existant (module VBA Import_Incidents_OGN).
# Conservées à l'identique pour que le rapport hebdomadaire reste exploitable
# avec les habitudes en place.
COLONNES_SUIVI = [
    "N°", "N° ticket", "Priorité", "Origine", "Nom de service", "Début",
    "Fin rétablissement", "Fin réparation", "Durée rétablissement (hh:mn)",
    "Durée rétablissement (mn)", "Durée réparation (hh:mn)",
    "Durée réparation (mn)", "Description", "Cause de l'incident",
    "Actions correctives", "TMC", "Statut", "Observation", "SLA",
    "Exclusion", "RI",
]


# ======================================================================
#  Utilitaires de période
# ======================================================================

def _dt(valeur: str | None) -> datetime | None:
    if not valeur:
        return None
    txt = str(valeur).replace("T", " ").strip()
    for motif in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(txt, motif)
        except ValueError:
            continue
    return None


def _duree_minutes(debut: str | None, fin: str | None) -> int | None:
    d, f = _dt(debut), _dt(fin)
    if not d or not f or f < d:
        return None
    return int((f - d).total_seconds() // 60)


def _hhmm(minutes: int | None) -> str:
    if minutes is None:
        return ""
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def bornes_shift(moment: datetime | None = None, params: dict | None = None
                 ) -> tuple[datetime, datetime, int]:
    """
    Détermine le shift contenant `moment`.

    Renvoie (début, fin, numéro). Le nombre de shifts et l'heure de départ du
    premier viennent du paramétrage, jamais du code.
    """
    params = params or tous_parametres()
    moment = moment or datetime.now()

    nb = int(params.get("rapport.shifts_par_jour", 3) or 3)
    duree = 24 / nb

    depart = str(params.get("rapport.heure_debut_shift1", "06:00"))
    try:
        h0, m0 = (int(x) for x in depart.split(":"))
    except ValueError:
        h0, m0 = 6, 0

    origine = moment.replace(hour=h0, minute=m0, second=0, microsecond=0)
    if moment < origine:
        origine -= timedelta(days=1)

    ecoule = (moment - origine).total_seconds() / 3600
    index = int(ecoule // duree)

    debut = origine + timedelta(hours=index * duree)
    return debut, debut + timedelta(hours=duree), index + 1


def bornes_semaine(jour: date | None = None, params: dict | None = None
                   ) -> tuple[date, date]:
    """Semaine calendaire, avec premier jour configurable (1 = lundi)."""
    params = params or tous_parametres()
    jour = jour or date.today()
    premier = int(params.get("rapport.jour_debut_semaine", 1) or 1)
    decalage = (jour.isoweekday() - premier) % 7
    debut = jour - timedelta(days=decalage)
    return debut, debut + timedelta(days=6)


def bornes_mois(jour: date | None = None) -> tuple[date, date]:
    jour = jour or date.today()
    debut = jour.replace(day=1)
    fin = (debut.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    return debut, fin


# ======================================================================
#  Extraction
# ======================================================================

def incidents_periode(debut: datetime | str, fin: datetime | str) -> list[dict]:
    """
    Incidents concernés par une période.

    Un incident est retenu s'il a démarré, s'est terminé, ou a fait l'objet
    d'une communication pendant la fenêtre. Sans cette troisième condition, un
    incident de longue durée disparaîtrait des rapports intermédiaires alors
    qu'il est justement celui qu'il faut signaler à l'équipe suivante.
    """
    d = debut.strftime("%Y-%m-%d %H:%M") if isinstance(debut, datetime) else str(debut)
    f = fin.strftime("%Y-%m-%d %H:%M") if isinstance(fin, datetime) else str(fin)

    conn = get_conn()
    try:
        lignes = [dict(r) for r in conn.execute(
            """SELECT i.*,
                      (SELECT COUNT(*) FROM incident_communications c
                        WHERE c.incident_id = i.id) AS nb_communications,
                      (SELECT GROUP_CONCAT(d.nom || ' / ' || s.nom, ' ; ')
                         FROM incident_services x
                         JOIN services s ON s.id = x.service_id
                         JOIN domaines d ON d.id = s.domaine_id
                        WHERE x.incident_id = i.id) AS services_nom,
                      (SELECT d.nom FROM incident_services x
                         JOIN services s ON s.id = x.service_id
                         JOIN domaines d ON d.id = s.domaine_id
                        WHERE x.incident_id = i.id LIMIT 1) AS domaine
               FROM incidents i
               WHERE (i.date_debut BETWEEN ? AND ?)
                  OR (i.date_fin   BETWEEN ? AND ?)
                  OR EXISTS (SELECT 1 FROM incident_communications c
                              WHERE c.incident_id = i.id
                                AND c.envoye_at BETWEEN ? AND ?)
               ORDER BY i.date_debut""",
            (d, f, d, f, d, f))]
    finally:
        conn.close()

    for l in lignes:
        l["attributs_specifiques"] = lire_attributs(l["attributs_specifiques"])
        l["duree_minutes"] = _duree_minutes(l["date_debut"], l["date_fin"])
        l["duree_hhmm"] = _hhmm(l["duree_minutes"])
        l["en_cours"] = not l["date_fin"]
        l["perimetre"] = l["services_nom"] or l["perimetre_libre"] or ""
    return lignes


# ======================================================================
#  Rapport de fin de shift
# ======================================================================

def rapport_shift(moment: datetime | None = None) -> dict:
    params = tous_parametres()
    debut, fin, numero = bornes_shift(moment, params)
    incidents = incidents_periode(debut, fin)

    seuil = str(params.get("rapport.seuil_incident_majeur", "P1"))
    rang = {"P1": 1, "P2": 2, "P3": 3}
    seuil_rang = rang.get(seuil, 1)

    en_cours = [i for i in incidents if i["en_cours"]]
    clos = [i for i in incidents if not i["en_cours"]]
    majeurs = [i for i in incidents if rang.get(i["priorite"] or "", 9) <= seuil_rang]

    # Ce qui demande une attention de l'équipe suivante : incidents encore
    # ouverts, et incidents clos dont l'action reste en attente.
    a_suivre = en_cours + [
        i for i in clos
        if (i.get("action") or "").lower().startswith("en attente")
    ]

    return {
        "type": "shift",
        "titre": f"Rapport de fin de shift — shift {numero}",
        "debut": debut.strftime("%d/%m/%Y %H:%M"),
        "fin": fin.strftime("%d/%m/%Y %H:%M"),
        "total": len(incidents),
        "en_cours": len(en_cours),
        "clos": len(clos),
        "majeurs": len(majeurs),
        "incidents": incidents,
        "a_suivre": a_suivre,
        "seuil_majeur": seuil,
    }


# ======================================================================
#  Rapport hebdomadaire
# ======================================================================

def rapport_hebdomadaire(jour: date | None = None) -> dict:
    params = tous_parametres()
    debut, fin = bornes_semaine(jour, params)
    d = datetime.combine(debut, datetime.min.time())
    f = datetime.combine(fin, datetime.max.time())
    incidents = incidents_periode(d, f)

    par_priorite: dict = {}
    par_domaine: dict = {}
    par_jour: dict = {}
    durees = []

    for i in incidents:
        par_priorite[i["priorite"] or "—"] = par_priorite.get(i["priorite"] or "—", 0) + 1
        dom = i.get("domaine") or "Non catalogué"
        par_domaine[dom] = par_domaine.get(dom, 0) + 1
        dt = _dt(i["date_debut"])
        if dt:
            cle = dt.strftime("%Y-%m-%d")
            par_jour[cle] = par_jour.get(cle, 0) + 1
        if i["duree_minutes"] is not None:
            durees.append(i["duree_minutes"])

    ouverts = [i for i in incidents if i["en_cours"]]
    regularises = [i for i in incidents if i["statut_documentaire"] == "regularise"]

    return {
        "type": "hebdomadaire",
        "titre": f"Rapport hebdomadaire — semaine du {debut.strftime('%d/%m/%Y')}",
        "debut": debut.strftime("%d/%m/%Y"),
        "fin": fin.strftime("%d/%m/%Y"),
        "total": len(incidents),
        "clos": len(incidents) - len(ouverts),
        "ouverts_fin_periode": len(ouverts),
        "regularisations": len(regularises),
        "duree_moyenne": round(sum(durees) / len(durees)) if durees else None,
        "duree_max": max(durees) if durees else None,
        "par_priorite": par_priorite,
        "par_domaine": dict(sorted(par_domaine.items(), key=lambda x: -x[1])),
        "par_jour": par_jour,
        "incidents": incidents,
        "liste_ouverts": ouverts,
        "liste_regularisations": regularises,
    }


# ======================================================================
#  Rapport mensuel
# ======================================================================

def rapport_mensuel(jour: date | None = None) -> dict:
    params = tous_parametres()
    debut, fin = bornes_mois(jour)
    d = datetime.combine(debut, datetime.min.time())
    f = datetime.combine(fin, datetime.max.time())
    incidents = incidents_periode(d, f)

    seuil = str(params.get("rapport.seuil_incident_majeur", "P1"))
    rang = {"P1": 1, "P2": 2, "P3": 3}
    seuil_rang = rang.get(seuil, 1)

    par_type: dict = {}
    par_priorite: dict = {}
    par_domaine: dict = {}
    par_semaine: dict = {}
    durees = []

    for i in incidents:
        par_type[i["type_incident"]] = par_type.get(i["type_incident"], 0) + 1
        par_priorite[i["priorite"] or "—"] = par_priorite.get(i["priorite"] or "—", 0) + 1
        dom = i.get("domaine") or "Non catalogué"
        par_domaine[dom] = par_domaine.get(dom, 0) + 1
        dt = _dt(i["date_debut"])
        if dt:
            sem = f"S{dt.isocalendar()[1]}"
            par_semaine[sem] = par_semaine.get(sem, 0) + 1
        if i["duree_minutes"] is not None:
            durees.append(i["duree_minutes"])

    majeurs = sorted(
        [i for i in incidents if rang.get(i["priorite"] or "", 9) <= seuil_rang],
        key=lambda x: x["duree_minutes"] or 0, reverse=True)

    # Tendance : comparaison au mois précédent, pour situer le volume.
    mois_precedent = (debut - timedelta(days=1)).replace(day=1)
    fin_precedent = debut - timedelta(days=1)
    precedents = incidents_periode(
        datetime.combine(mois_precedent, datetime.min.time()),
        datetime.combine(fin_precedent, datetime.max.time()))

    ecart = None
    if precedents:
        ecart = round((len(incidents) - len(precedents)) / len(precedents) * 100)

    return {
        "type": "mensuel",
        "titre": f"Rapport mensuel — {debut.strftime('%m/%Y')}",
        "debut": debut.strftime("%d/%m/%Y"),
        "fin": fin.strftime("%d/%m/%Y"),
        "total": len(incidents),
        "total_mois_precedent": len(precedents),
        "evolution_pct": ecart,
        "clos": sum(1 for i in incidents if not i["en_cours"]),
        "ouverts_fin_periode": sum(1 for i in incidents if i["en_cours"]),
        "duree_moyenne": round(sum(durees) / len(durees)) if durees else None,
        "duree_max": max(durees) if durees else None,
        "par_type": par_type,
        "par_priorite": par_priorite,
        "par_domaine": dict(sorted(par_domaine.items(), key=lambda x: -x[1])),
        "par_semaine": dict(sorted(par_semaine.items())),
        "incidents_majeurs": majeurs[:20],
        "nb_majeurs": len(majeurs),
        "seuil_majeur": seuil,
        "incidents": incidents,
    }


# ======================================================================
#  Export Excel
# ======================================================================

def exporter_excel(rapport: dict, chemin: str) -> str:
    """
    Produit un classeur reprenant les colonnes du suivi existant.

    Les totaux sont écrits en formules (`SUM`, `COUNTIF`) et non en valeurs
    calculées : le classeur reste juste si le superviseur ajoute ou corrige
    une ligne à la main.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Charte du classeur : sobre, lisible à l'impression.
    POLICE = "Arial"
    titre_f = Font(name=POLICE, size=14, bold=True)
    entete_f = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
    normal_f = Font(name=POLICE, size=10)
    gras_f = Font(name=POLICE, size=10, bold=True)
    remplissage = PatternFill("solid", fgColor="404040")
    bord = Border(*[Side(style="thin", color="D0D0D0")] * 4)

    # ---------------- Feuille de synthèse ----------------
    ws = wb.active
    ws.title = "Synthèse"
    ws["A1"] = rapport["titre"]
    ws["A1"].font = titre_f
    ws["A2"] = f"Période du {rapport['debut']} au {rapport['fin']}"
    ws["A2"].font = Font(name=POLICE, size=10, italic=True, color="666666")

    ligne = 4
    indicateurs = [("Total incidents", rapport["total"])]
    if "clos" in rapport:
        indicateurs.append(("Clôturés", rapport["clos"]))
    for cle, libelle in (("en_cours", "Encore en cours"),
                         ("ouverts_fin_periode", "Encore ouverts en fin de période"),
                         ("majeurs", "Incidents majeurs"),
                         ("nb_majeurs", "Incidents majeurs"),
                         ("regularisations", "Régularisations")):
        if cle in rapport:
            indicateurs.append((libelle, rapport[cle]))
    if rapport.get("duree_moyenne") is not None:
        indicateurs.append(("Durée moyenne (mn)", rapport["duree_moyenne"]))
        indicateurs.append(("Durée maximale (mn)", rapport["duree_max"]))
    if rapport.get("evolution_pct") is not None:
        indicateurs.append(("Évolution vs mois précédent (%)", rapport["evolution_pct"]))

    for libelle, valeur in indicateurs:
        ws.cell(ligne, 1, libelle).font = gras_f
        ws.cell(ligne, 2, valeur).font = normal_f
        ligne += 1

    # Répartitions
    for cle, titre in (("par_priorite", "Par priorité"),
                       ("par_type", "Par type"),
                       ("par_domaine", "Par domaine")):
        if not rapport.get(cle):
            continue
        ligne += 1
        ws.cell(ligne, 1, titre).font = gras_f
        ligne += 1
        for k, v in list(rapport[cle].items())[:15]:
            ws.cell(ligne, 1, str(k)).font = normal_f
            ws.cell(ligne, 2, v).font = normal_f
            ligne += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14

    # ---------------- Feuille détail ----------------
    wd = wb.create_sheet("Incidents")
    for c, entete in enumerate(COLONNES_SUIVI, 1):
        cel = wd.cell(1, c, entete)
        cel.font = entete_f
        cel.fill = remplissage
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = bord
    wd.row_dimensions[1].height = 30

    for n, i in enumerate(rapport["incidents"], 1):
        L = n + 1
        valeurs = [
            n,
            i["reference_externe"],
            i["priorite"] or "",
            "Supervision",
            i["perimetre"],
            _dt(i["date_debut"]),
            _dt(i["date_fin"]),
            "",                                  # fin réparation : saisie manuelle
            i["duree_hhmm"],
            i["duree_minutes"] if i["duree_minutes"] is not None else "",
            "", "",                              # durées réparation : manuelles
            i["description"],
            i.get("cause") or "",
            i.get("action") or "",
            i["attributs_specifiques"].get("tmc", ""),
            "En cours" if i["en_cours"] else "Résolu",
            i.get("observation") or "",
            "", "", "",                          # SLA / Exclusion / RI : manuels
        ]
        for c, v in enumerate(valeurs, 1):
            cel = wd.cell(L, c, v)
            cel.font = normal_f
            cel.border = bord
            cel.alignment = Alignment(vertical="top", wrap_text=(c in (13, 14, 15, 18)))
            if c in (6, 7) and isinstance(v, datetime):
                cel.number_format = "DD/MM/YYYY HH:MM"

    # Ligne de totaux, en formules pour rester juste après édition manuelle
    if rapport["incidents"]:
        derniere = len(rapport["incidents"]) + 1
        T = derniere + 2
        wd.cell(T, 1, "TOTAUX").font = gras_f
        wd.cell(T, 2, f'=COUNTA(B2:B{derniere})').font = gras_f
        wd.cell(T, 10, f'=SUM(J2:J{derniere})').font = gras_f
        wd.cell(T + 1, 1, "Durée moyenne (mn)").font = gras_f
        wd.cell(T + 1, 10, f'=IFERROR(AVERAGE(J2:J{derniere}),0)').font = gras_f
        wd.cell(T + 2, 1, "Incidents en cours").font = gras_f
        wd.cell(T + 2, 10, f'=COUNTIF(Q2:Q{derniere},"En cours")').font = gras_f

    largeurs = [5, 13, 8, 12, 30, 17, 17, 17, 12, 11, 12, 11,
                40, 32, 32, 12, 11, 22, 8, 10, 8]
    for c, w in enumerate(largeurs, 1):
        wd.column_dimensions[get_column_letter(c)].width = w
    wd.freeze_panes = "A2"
    wd.auto_filter.ref = f"A1:U{max(2, len(rapport['incidents']) + 1)}"

    wb.save(chemin)
    return chemin


# ======================================================================
#  Rendu HTML (aperçu et impression)
# ======================================================================

def rendre_html(rapport: dict) -> str:
    """
    Rendu du rapport, généré par le serveur.

    Charte volontairement neutre et imprimable : ce document circule par
    e-mail et sur papier, il ne suit donc ni le thème de l'application ni
    la charte des avis d'incident.
    """
    import html as H

    def bloc(titre: str, contenu: str) -> str:
        return (f'<div class="bloc"><div class="bloc-titre">{H.escape(titre)}</div>'
                f'{contenu}</div>')

    def repartition(donnees: dict) -> str:
        if not donnees:
            return '<p class="neant">Aucune donnée</p>'
        total = sum(donnees.values()) or 1
        lignes = "".join(
            f'<tr><td>{H.escape(str(k))}</td><td class="n">{v}</td>'
            f'<td class="b"><span style="width:{round(v / total * 100)}%"></span></td></tr>'
            for k, v in list(donnees.items())[:12])
        return f'<table class="rep">{lignes}</table>'

    def table_incidents(liste: list, colonnes: list) -> str:
        if not liste:
            return '<p class="neant">Aucun incident</p>'
        entetes = "".join(f"<th>{H.escape(t)}</th>" for _, t in colonnes)
        corps = ""
        for i in liste:
            cells = ""
            for cle, _ in colonnes:
                v = i.get(cle) or ""
                if cle == "statut":
                    v = "En cours" if i["en_cours"] else "Résolu"
                cells += f"<td>{H.escape(str(v))}</td>"
            corps += f"<tr>{cells}</tr>"
        return f"<table class='det'><thead><tr>{entetes}</tr></thead><tbody>{corps}</tbody></table>"

    indicateurs = [("Total", rapport["total"])]
    for cle, lib in (("en_cours", "En cours"), ("clos", "Clôturés"),
                     ("ouverts_fin_periode", "Ouverts en fin de période"),
                     ("majeurs", "Majeurs"), ("nb_majeurs", "Majeurs"),
                     ("regularisations", "Régularisations")):
        if cle in rapport:
            indicateurs.append((lib, rapport[cle]))
    if rapport.get("duree_moyenne") is not None:
        indicateurs.append(("Durée moyenne", _hhmm(rapport["duree_moyenne"])))

    cartes = "".join(
        f'<div class="carte"><div class="v">{H.escape(str(v))}</div>'
        f'<div class="l">{H.escape(l)}</div></div>' for l, v in indicateurs)

    corps = f'<div class="cartes">{cartes}</div>'

    if rapport["type"] == "shift":
        corps += bloc("À suivre par l'équipe suivante", table_incidents(
            rapport["a_suivre"],
            [("reference_externe", "Référence"), ("priorite", "Prio"),
             ("perimetre", "Service"), ("statut", "Statut"),
             ("action", "Action en cours"), ("observation", "Observation")]))
        corps += bloc("Tous les incidents du shift", table_incidents(
            rapport["incidents"],
            [("reference_externe", "Référence"), ("priorite", "Prio"),
             ("perimetre", "Service"), ("date_debut", "Début"),
             ("duree_hhmm", "Durée"), ("statut", "Statut")]))

    elif rapport["type"] == "hebdomadaire":
        corps += ('<div class="deux">'
                  + bloc("Par priorité", repartition(rapport["par_priorite"]))
                  + bloc("Par domaine", repartition(rapport["par_domaine"]))
                  + "</div>")
        if rapport["liste_ouverts"]:
            corps += bloc("Encore ouverts en fin de semaine", table_incidents(
                rapport["liste_ouverts"],
                [("reference_externe", "Référence"), ("priorite", "Prio"),
                 ("perimetre", "Service"), ("date_debut", "Début"),
                 ("action", "Action")]))
        if rapport["liste_regularisations"]:
            corps += bloc("Régularisations de la semaine", table_incidents(
                rapport["liste_regularisations"],
                [("reference_externe", "Référence"), ("perimetre", "Service"),
                 ("date_debut", "Début"), ("duree_hhmm", "Durée")]))
        corps += bloc("Détail des incidents", table_incidents(
            rapport["incidents"],
            [("reference_externe", "Référence"), ("priorite", "Prio"),
             ("perimetre", "Service"), ("date_debut", "Début"),
             ("duree_hhmm", "Durée"), ("cause", "Cause"), ("statut", "Statut")]))

    else:  # mensuel
        if rapport.get("evolution_pct") is not None:
            signe = "+" if rapport["evolution_pct"] > 0 else ""
            corps += (f'<p class="tendance">Évolution du volume : '
                      f'<b>{signe}{rapport["evolution_pct"]} %</b> '
                      f'({rapport["total_mois_precedent"]} le mois précédent)</p>')
        corps += ('<div class="deux">'
                  + bloc("Par type", repartition(rapport["par_type"]))
                  + bloc("Par priorité", repartition(rapport["par_priorite"]))
                  + "</div>")
        corps += ('<div class="deux">'
                  + bloc("Par domaine de service", repartition(rapport["par_domaine"]))
                  + bloc("Répartition hebdomadaire", repartition(rapport["par_semaine"]))
                  + "</div>")
        corps += bloc(f"Incidents majeurs ({rapport['seuil_majeur']} et au-dessus)",
                      table_incidents(rapport["incidents_majeurs"],
                      [("reference_externe", "Référence"), ("priorite", "Prio"),
                       ("perimetre", "Service"), ("date_debut", "Début"),
                       ("duree_hhmm", "Durée"), ("cause", "Cause")]))

    return f"""<!DOCTYPE html><html lang="fr"><head><meta charset="utf-8"><style>
* {{ box-sizing: border-box; }}
body {{ font-family: Arial, Helvetica, sans-serif; font-size: 12px; color: #1A1A1A;
        background: #fff; margin: 0; padding: 22px; }}
h1 {{ font-size: 17px; margin: 0 0 4px; }}
.periode {{ color: #666; font-size: 12px; margin-bottom: 18px; }}
.cartes {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 20px; }}
.carte {{ flex: 1; min-width: 96px; border: 1px solid #DDD; border-radius: 5px;
          padding: 11px 13px; }}
.carte .v {{ font-size: 21px; font-weight: 700; }}
.carte .l {{ font-size: 10.5px; color: #666; margin-top: 2px; }}
.bloc {{ margin-bottom: 20px; }}
.bloc-titre {{ font-size: 11px; font-weight: 700; text-transform: uppercase;
               letter-spacing: .08em; color: #444; padding-bottom: 6px;
               margin-bottom: 10px; border-bottom: 2px solid #333; }}
.deux {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 11.5px; }}
.rep td {{ padding: 4px 6px; border-bottom: 1px solid #EEE; }}
.rep td.n {{ text-align: right; width: 46px; font-weight: 700; }}
.rep td.b {{ width: 40%; }}
.rep td.b span {{ display: block; height: 7px; background: #555; border-radius: 2px; }}
.det th {{ text-align: left; padding: 6px; background: #F2F2F2;
           border-bottom: 2px solid #CCC; font-size: 10.5px; }}
.det td {{ padding: 6px; border-bottom: 1px solid #EEE; vertical-align: top; }}
.neant {{ color: #999; font-style: italic; font-size: 11.5px; }}
.tendance {{ padding: 9px 12px; background: #F7F7F7; border-left: 3px solid #555;
             margin-bottom: 18px; }}
@media print {{ body {{ padding: 0; }} .bloc {{ page-break-inside: avoid; }} }}
</style></head><body>
<h1>{H.escape(rapport['titre'])}</h1>
<div class="periode">Période du {H.escape(rapport['debut'])} au {H.escape(rapport['fin'])}</div>
{corps}
</body></html>"""
